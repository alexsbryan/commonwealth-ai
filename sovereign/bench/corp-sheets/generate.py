#!/usr/bin/env python3
"""Generate a synthetic corporate-spreadsheet corpus for exercising the
described-asset / XLSX column-aware enrichment substrate.

Why synthetic: the Enron maildir corpus is body-only (0 attachments), so
it cannot exercise the attachment path. This generates the *common*
corporate spreadsheet shapes — org charts, operating budgets, a DCF
model, a vendor master, a headcount roster — with a realistic amount of
corporate NOISE (merged banner rows, multi-row headers, subtotal/total
rows, footnotes, blank separators, N/A / TBD cells, free-text Notes
columns) and a *planted cast* of person + organization entities that
recur across sheets with inconsistent surface forms ("Katherine Chen" /
"Chen, Katherine" / "K. Chen" / "kchen@acme.com"). That recurrence is
deliberate: it exercises both the column-aware extractor (typed-header →
entity) and the reconciliation layer (name-variant + corporate-suffix
collapse) at once.

Outputs:
  - workbooks under  ~/.svrnmesh/corpora-staging/corp-sheets/*.xlsx
  - gold manifest    sovereign/bench/corp-sheets/ground_truth_entities.jsonl
    (canonical_id, entity_type, canonical_name, surface_forms[], the
     sheets each form appears in, and whether it sits in a typed
     [column-aware-visible] column or noise-only)

Reproducible: deterministic content, no randomness. Re-run to regenerate.
"""
import json
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = Path.home() / ".sovereign" / "corpora-staging" / "corp-sheets"
GOLD = Path(__file__).resolve().parent / "ground_truth_entities.jsonl"

# ── Planted cast ─────────────────────────────────────────────────
# Each entity: canonical_id -> (type, canonical_name, {form: note}).
# Forms are the inconsistent surface variants we will scatter across
# sheets. `surface_forms` in the gold = the keys.
PEOPLE = {
    "person-katherine-chen": ("Katherine Chen",
        ["Katherine Chen", "Chen, Katherine", "K. Chen", "Katherine M. Chen", "kchen@acme.com"]),
    "person-marcus-webb": ("Marcus Webb",
        ["Marcus Webb", "Webb, Marcus", "M. Webb", "marcus.webb@acme.com"]),
    "person-priya-anand": ("Priya Anand",
        ["Priya Anand", "Anand, Priya", "P. Anand", "priya.anand@acme.com"]),
    "person-james-obrien": ("James O'Brien",
        ["James O'Brien", "O'Brien, James", "Jim O'Brien", "jobrien@acme.com"]),
    "person-daniel-ruiz": ("Daniel Ruiz",
        ["Daniel Ruiz", "Ruiz, Daniel", "D. Ruiz", "daniel.ruiz@acme.com"]),
    "person-sofia-larsen": ("Sofia Larsen",
        ["Sofia Larsen", "Larsen, Sofia", "S. Larsen", "sofia.larsen@acme.com"]),
    "person-omar-haddad": ("Omar Haddad",
        ["Omar Haddad", "Haddad, Omar", "O. Haddad", "omar.haddad@acme.com"]),
    # GENERALIZATION PROBE — appears ONLY under cryptic headers
    # ("DRI", "Resp. Party") the keyword map does not enumerate. Embed
    # classification must catch these; keyword mode will miss them.
    "person-tobias-vance": ("Tobias Vance",
        ["Tobias Vance", "Vance, Tobias", "T. Vance"]),
}
ORGS = {
    "org-dynacorp": ("Dynacorp Industries Inc.",
        ["Dynacorp Industries Inc.", "Dynacorp", "Dynacorp Inc.", "Dynacorp Industries"]),
    "org-meridian": ("Meridian Capital Partners LLC",
        ["Meridian Capital Partners LLC", "Meridian Capital", "Meridian", "Meridian Capital Partners"]),
    "org-northwind": ("Northwind Logistics Corp.",
        ["Northwind Logistics Corp.", "Northwind", "Northwind Logistics"]),
    "org-salesforce": ("Salesforce.com, Inc.",
        ["Salesforce.com, Inc.", "Salesforce"]),
    "org-aws": ("Amazon Web Services, Inc.",
        ["Amazon Web Services, Inc.", "AWS", "Amazon Web Services"]),
    "org-globex": ("Globex Corporation",
        ["Globex Corporation", "Globex", "Globex Corp."]),
    # GENERALIZATION PROBE — appears ONLY under a cryptic "Cpty" header.
    "org-vanguard": ("Vanguard Holdings LLC",
        ["Vanguard Holdings LLC", "Vanguard Holdings", "Vanguard"]),
}

# track which forms we actually emit + in which sheets / typed-or-noise
emitted = {}  # form -> {"canonical_id", "sheets": set, "typed": bool}

def emit(form, canonical_id, sheet, typed):
    rec = emitted.setdefault(form, {"canonical_id": canonical_id, "sheets": set(), "typed": False})
    rec["sheets"].add(sheet)
    rec["typed"] = rec["typed"] or typed

# style helpers ----------------------------------------------------
BANNER = Font(bold=True, size=14, color="FFFFFF")
BANNER_FILL = PatternFill("solid", fgColor="1F3864")
HDR = Font(bold=True)
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
FOOT = Font(italic=True, size=9, color="808080")

def banner(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text); c.font = BANNER; c.fill = BANNER_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

def headerrow(ws, r, cols):
    for i, h in enumerate(cols, 1):
        c = ws.cell(r, i, h); c.font = HDR; c.fill = HDR_FILL

def footnote(ws, r, text):
    c = ws.cell(r, 1, text); c.font = FOOT

# ── Sheet builders ───────────────────────────────────────────────

def org_chart(path):
    wb = Workbook(); ws = wb.active; ws.title = "Org Chart"
    banner(ws, "Acme Corporation — Organizational Chart (FY2024)", 5)
    ws.append([])  # row 2 blank
    headerrow(ws, 3, ["Employee", "Title", "Reports To", "Department", "Email"])
    rows = [
        ("James O'Brien", "Chief Executive Officer", "", "Executive", "jobrien@acme.com"),
        ("Katherine Chen", "Chief Financial Officer", "O'Brien, James", "Finance", "kchen@acme.com"),
        ("Marcus Webb", "VP, Engineering", "O'Brien, James", "Engineering", "marcus.webb@acme.com"),
        ("Priya Anand", "Controller", "Chen, Katherine", "Finance", "priya.anand@acme.com"),
        ("Daniel Ruiz", "Director, Platform", "Webb, Marcus", "Engineering", "daniel.ruiz@acme.com"),
        ("Sofia Larsen", "Head of Procurement", "Chen, Katherine", "Operations", "sofia.larsen@acme.com"),
    ]
    r = 4
    for name, title, mgr, dept, email in rows:
        ws.append([name, title, mgr, dept, email]);
        emit(name, pid(name), "org_chart", True)        # Employee col → typed Person
        emit(email, pid(name), "org_chart", True)        # Email col → typed
        if mgr: emit(mgr, pid_from_lastfirst(mgr), "org_chart", False)  # Reports To → not a typed header
        r += 1
    ws.append([]); r += 1
    # noisy sub-section with its own mini-header
    c = ws.cell(r, 1, "Contractors & Advisors (non-FTE)"); c.font = Font(bold=True, italic=True); r += 1
    headerrow(ws, r, ["Name", "Role", "Firm", "", "Contact"]); r += 1
    ws.cell(r, 1, "Omar Haddad"); ws.cell(r, 2, "Security Advisor"); ws.cell(r, 3, "Meridian Capital"); ws.cell(r, 5, "omar.haddad@acme.com")
    emit("Omar Haddad", "person-omar-haddad", "org_chart", True)   # Name col → typed
    emit("omar.haddad@acme.com", "person-omar-haddad", "org_chart", True)
    emit("Meridian Capital", "org-meridian", "org_chart", False)   # Firm not a typed header
    r += 1
    ws.append([]); r += 1
    footnote(ws, r, "* Org chart as of Q1 2024; pending reorg. Source: People Ops.")
    widths(ws, [22, 26, 18, 16, 26]); wb.save(path)

def budget(path):
    wb = Workbook()
    ws = wb.active; ws.title = "Opex"
    banner(ws, "FY2024 Operating Budget — CONFIDENTIAL (Draft v3)", 7)
    ws.append([]);
    # multi-row group header
    ws.cell(3, 5, "FY2024 ($000s)").font = HDR; ws.merge_cells("E3:H3")
    headerrow(ws, 4, ["Department", "Line Item", "Owner", "Vendor", "Q1", "Q2", "Q3", "Q4"])
    rows = [
        ("Engineering", "Cloud infrastructure", "M. Webb", "Amazon Web Services", 420, 440, 455, 470),
        ("Engineering", "CRM licenses", "D. Ruiz", "Salesforce", 80, 80, 85, 85),
        ("Engineering", "Subtotal", "", "", 500, 520, 540, 555),
        ("Finance", "Audit & advisory", "K. Chen", "Meridian Capital Partners", 120, 0, 0, 130),
        ("Finance", "Treasury services", "P. Anand", "Dynacorp", 35, 35, 35, 35),
        ("Operations", "Logistics", "S. Larsen", "Northwind Logistics", 210, 215, 220, 225),
        ("Operations", "Logistics — peak", "S. Larsen", "Northwind Logistics Corp.", 0, 0, 60, 90),
    ]
    r = 5
    for dept, item, owner, vendor, *q in rows:
        ws.append([dept, item, owner, vendor, *q])
        if owner: emit(owner, pid_from_initial(owner), "budget_opex", True)   # Owner → typed Person
        if vendor: emit(vendor, oid(vendor), "budget_opex", True)             # Vendor → typed Org
        r += 1
    ws.append(["TOTAL", "", "", "", 1000, 1090, 1135, 1190]); r += 1
    ws.append([]); r += 1
    footnote(ws, r, "Source: FP&A model BUD-FY24-v3. * Q2 audit deferred. Owners per cost-center matrix.")
    widths(ws, [16, 22, 12, 28, 8, 8, 8, 8])
    # Capex sheet (more noise + reuse)
    ws2 = wb.create_sheet("Capex")
    banner(ws2, "FY2024 Capital Expenditure", 5)
    ws2.append([])
    headerrow(ws2, 3, ["Project", "Sponsor", "Counterparty", "Budget", "Status"])
    caprows = [
        ("Data center expansion", "Marcus Webb", "Dynacorp Industries Inc.", 2400, "Approved"),
        ("ERP migration", "Katherine Chen", "Globex Corporation", 1800, "Pending"),
        ("Fleet refresh", "Sofia Larsen", "Northwind", 600, "TBD"),
    ]
    for proj, sponsor, cpty, bud, status in caprows:
        ws2.append([proj, sponsor, cpty, bud, status])
        emit(sponsor, pid(sponsor), "budget_capex", False)     # Sponsor not a typed header
        emit(cpty, oid(cpty), "budget_capex", True)            # Counterparty → typed Org
    widths(ws2, [24, 18, 26, 10, 12]); wb.save(path)

def dcf(path):
    wb = Workbook()
    ws = wb.active; ws.title = "DCF"
    banner(ws, "Project Falcon — Discounted Cash Flow (Draft, Confidential)", 7)
    ws.append([])
    # assumptions block (label/value pairs — pure noise to column-aware)
    asmp = [("WACC", "9.5%"), ("Terminal growth", "2.5%"), ("Tax rate", "21%"),
            ("Forecast horizon", "5 yrs"), ("Valuation date", "2024-03-31")]
    ws.cell(3, 1, "Key Assumptions").font = HDR
    r = 4
    for k, v in asmp:
        ws.cell(r, 1, k); ws.cell(r, 2, v); r += 1
    r += 1
    # projection table — years as columns
    headerrow(ws, r, ["Line ($mm)", "2024E", "2025E", "2026E", "2027E", "2028E"]); r += 1
    proj = [("Revenue", 320, 360, 405, 455, 510), ("EBITDA", 64, 79, 93, 109, 128),
            ("Unlevered FCF", 38, 50, 61, 74, 90), ("Discount factor", 0.91, 0.83, 0.76, 0.69, 0.63)]
    for line in proj:  # numbers are pure noise to entity extraction
        ws.cell(r, 1, line[0])
        for i, v in enumerate(line[1:], 2):
            ws.cell(r, i, v)
        r += 1
    r += 1
    footnote(ws, r, "Prepared by Katherine Chen, CFO. Reviewed by M. Webb (VP Eng). Target: Globex Corp.")
    emit("Katherine Chen", "person-katherine-chen", "dcf", False)  # footnote prose, not a column
    emit("M. Webb", "person-marcus-webb", "dcf", False)
    emit("Globex Corp.", "org-globex", "dcf", False)
    widths(ws, [16, 10, 10, 10, 10, 10])
    # Comps sheet — a typed Org column
    ws2 = wb.create_sheet("Comps")
    banner(ws2, "Comparable Companies", 4)
    ws2.append([])
    headerrow(ws2, 3, ["Company", "Ticker", "EV/EBITDA", "Note"])
    comps = [("Dynacorp Industries Inc.", "DYN", "11.2x", "primary comp"),
             ("Globex Corporation", "GLBX", "9.8x", "n/a"),
             ("Northwind Logistics Corp.", "NWL", "7.4x", "")]
    for co, tk, mult, note in comps:
        ws2.append([co, tk, mult, note])
        emit(co, oid(co), "dcf_comps", True)   # Company → typed Org
    widths(ws2, [28, 10, 12, 16]); wb.save(path)

def vendor_master(path):
    wb = Workbook(); ws = wb.active; ws.title = "Vendors"
    banner(ws, "Acme — Approved Vendor Master", 6)
    ws.append([])
    headerrow(ws, 3, ["Vendor", "Category", "Primary Contact", "Email", "Annual Spend", "Notes"])
    rows = [
        ("Amazon Web Services, Inc.", "Cloud", "Daniel Ruiz", "daniel.ruiz@acme.com", 1850, "MSA renewed 2024"),
        ("Salesforce.com, Inc.", "SaaS", "D. Ruiz", "", 340, "seats TBD"),
        ("Meridian Capital Partners LLC", "Advisory", "Katherine Chen", "kchen@acme.com", 250, "audit + M&A"),
        ("Dynacorp Inc.", "Components", "Priya Anand", "priya.anand@acme.com", 500, "see also Dynacorp Industries"),
        ("Northwind Logistics", "Freight", "Sofia Larsen", "N/A", 410, ""),
    ]
    for vendor, cat, contact, email, spend, note in rows:
        ws.append([vendor, cat, contact, email, spend, note])
        emit(vendor, oid(vendor), "vendor_master", True)     # Vendor → typed Org
        emit(contact, pid(contact_norm(contact)), "vendor_master", True)  # Primary Contact → typed Person
        if email and email != "N/A": emit(email, pid_from_email(email), "vendor_master", True)
    ws.append([]); footnote(ws, len(rows) + 5, "* Annual spend in $000s. Contacts are internal owners, not vendor reps.")
    widths(ws, [28, 12, 18, 26, 12, 28]); wb.save(path)

def headcount(path):
    wb = Workbook(); ws = wb.active; ws.title = "Headcount"
    banner(ws, "Headcount Roster — HR Confidential", 5)
    ws.append([])
    headerrow(ws, 3, ["Employee", "Manager", "Level", "Location", "Start Date"])
    rows = [
        ("Katherine M. Chen", "O'Brien, James", "E", "HQ", "2019-02-01"),
        ("Marcus Webb", "James O'Brien", "E", "HQ", "2020-06-15"),
        ("Daniel Ruiz", "Webb, Marcus", "M5", "Remote", "2021-09-01"),
        ("Sofia Larsen", "Chen, Katherine", "M6", "HQ", "2018-11-20"),
        ("Priya Anand", "Chen, Katherine", "M5", "HQ", "2022-01-10"),
        ("Omar Haddad", "", "C", "Remote", "2023-04-03"),
    ]
    for emp, mgr, lvl, loc, start in rows:
        ws.append([emp, mgr, lvl, loc, start])
        emit(emp, pid(emp), "headcount", True)          # Employee → typed Person
        if mgr: emit(mgr, pid_from_lastfirst(mgr), "headcount", False)  # Manager not a typed header
    widths(ws, [22, 18, 8, 12, 12]); wb.save(path)

def risk_register(path):
    # GENERALIZATION PROBE. Cryptic abbreviation headers the keyword map
    # does NOT enumerate: "Cpty" (counterparty/org), "DRI" (directly
    # responsible individual/person), "Resp. Party" (person). The planted
    # gold here — Vanguard Holdings + Tobias Vance — appears in NO other
    # sheet, so it is recovered only by the embed-centroid classifier
    # (header+values semantics), never by substring keyword matching.
    wb = Workbook(); ws = wb.active; ws.title = "Risk Register"
    banner(ws, "Enterprise Risk Register — Q1 (Confidential)", 5)
    ws.append([])
    headerrow(ws, 3, ["Risk ID", "Cpty", "DRI", "Resp. Party", "Status"])
    rows = [
        ("R-001", "Vanguard Holdings", "Tobias Vance", "Vance, Tobias", "Open"),
        ("R-002", "Vanguard Holdings LLC", "T. Vance", "Tobias Vance", "Mitigating"),
        ("R-003", "Dynacorp", "Priya Anand", "Anand, Priya", "Closed"),
        ("R-004", "Vanguard", "Tobias Vance", "Tobias Vance", "Open"),
    ]
    for rid, cpty, dri, resp, status in rows:
        ws.append([rid, cpty, dri, resp, status])
        emit(cpty, oid(cpty), "risk_register", True)   # Cpty → Org (cryptic)
        emit(dri, pid(dri), "risk_register", True)     # DRI → Person (cryptic)
        emit(resp, pid(resp), "risk_register", True)   # Resp. Party → Person (cryptic)
    widths(ws, [10, 22, 16, 16, 12]); wb.save(path)

# ── id resolvers (map a surface form back to its planted canonical) ─
def _index():
    idx = {}
    for cid, (_canon, forms) in {**PEOPLE, **ORGS}.items():
        for f in forms:
            idx[f.lower()] = cid
    return idx
IDX = _index()

def _resolve(form):
    cid = IDX.get(form.lower())
    if cid is None:
        raise SystemExit(f"PLANT ERROR: surface form {form!r} not in cast — add it to PEOPLE/ORGS")
    return cid
def pid(form): return _resolve(form)
def oid(form): return _resolve(form)
def pid_from_initial(form): return _resolve(form)
def pid_from_email(form): return _resolve(form)
def pid_from_lastfirst(form): return _resolve(form)
def contact_norm(form): return form

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    org_chart(OUT_DIR / "org_chart_2024.xlsx")
    budget(OUT_DIR / "operating_budget_fy24.xlsx")
    dcf(OUT_DIR / "project_falcon_dcf.xlsx")
    vendor_master(OUT_DIR / "vendor_master.xlsx")
    headcount(OUT_DIR / "headcount_roster.xlsx")
    risk_register(OUT_DIR / "risk_register.xlsx")

    # gold manifest: aggregate forms per canonical
    by_canon = {}
    for form, rec in emitted.items():
        c = by_canon.setdefault(rec["canonical_id"], {"forms": {}, "typed_any": False})
        c["forms"][form] = sorted(rec["sheets"])
        c["typed_any"] = c["typed_any"] or rec["typed"]
    GOLD.parent.mkdir(parents=True, exist_ok=True)
    with open(GOLD, "w") as f:
        for cid in sorted(by_canon):
            canon, _ = (PEOPLE if cid in PEOPLE else ORGS)[cid]
            etype = "person" if cid in PEOPLE else "organization"
            forms = sorted(by_canon[cid]["forms"])
            f.write(json.dumps({
                "canonical_id": cid,
                "entity_type": etype,
                "canonical_name": canon,
                "surface_forms": forms,
                "split": "train",
                "typed_column_visible": by_canon[cid]["typed_any"],
            }) + "\n")
    nfiles = len(list(OUT_DIR.glob("*.xlsx")))
    print(f"wrote {nfiles} workbooks -> {OUT_DIR}")
    print(f"gold: {len(by_canon)} canonical entities -> {GOLD}")
    print(f"total distinct surface forms planted: {len(emitted)}")

if __name__ == "__main__":
    main()
